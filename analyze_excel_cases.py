import argparse
import collections
import datetime
import json
from pathlib import Path

from openpyxl import load_workbook


def sample_rows(worksheet, max_rows=4, max_cols=8):
    rows = []
    upper_row = min(worksheet.max_row, 12)
    for row in worksheet.iter_rows(min_row=1, max_row=upper_row, values_only=True):
        values = []
        for value in row[:max_cols]:
            if value is None:
                values.append("")
            else:
                values.append(str(value).strip())
        if any(values):
            rows.append(values)
        if len(rows) >= max_rows:
            break
    return rows


def analyze_sheet(worksheet):
    style_ids = set()
    nonempty_cells = 0
    formula_cells = 0
    comment_cells = 0

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                nonempty_cells += 1
                style_ids.add(cell.style_id)
            if cell.data_type == "f":
                formula_cells += 1
            if cell.comment is not None:
                comment_cells += 1

    return {
        "sheet_name": worksheet.title,
        "max_row": worksheet.max_row,
        "max_column": worksheet.max_column,
        "nonempty_cells": nonempty_cells,
        "style_count": len(style_ids),
        "formula_cells": formula_cells,
        "comment_cells": comment_cells,
        "merged_range_count": len(worksheet.merged_cells.ranges),
        "freeze_panes": str(worksheet.freeze_panes) if worksheet.freeze_panes else "",
        "sample_rows": sample_rows(worksheet),
    }


def analyze_workbook(path):
    workbook = load_workbook(path, data_only=False)
    sheet_summaries = [analyze_sheet(worksheet) for worksheet in workbook.worksheets]

    return {
        "file_name": path.name,
        "sheet_count": len(workbook.sheetnames),
        "sheet_names": workbook.sheetnames,
        "formula_cells": sum(item["formula_cells"] for item in sheet_summaries),
        "comment_cells": sum(item["comment_cells"] for item in sheet_summaries),
        "merged_range_count": sum(item["merged_range_count"] for item in sheet_summaries),
        "freeze_panes": [
            f"{item['sheet_name']}:{item['freeze_panes']}"
            for item in sheet_summaries
            if item["freeze_panes"]
        ],
        "sheets": sheet_summaries,
    }


def analyze_cases(case_dir):
    workbook_paths = sorted(case_dir.glob("*.xlsx"))
    recurring_sheets = collections.defaultdict(list)
    workbooks = []

    for workbook_path in workbook_paths:
        workbook_summary = analyze_workbook(workbook_path)
        workbooks.append(workbook_summary)
        for sheet_name in workbook_summary["sheet_names"]:
            recurring_sheets[sheet_name].append(workbook_summary["file_name"])

    recurring_sheet_names = []
    for sheet_name, file_names in recurring_sheets.items():
        recurring_sheet_names.append({
            "sheet_name": sheet_name,
            "count": len(file_names),
            "workbooks": sorted(file_names),
        })
    recurring_sheet_names.sort(key=lambda item: (-item["count"], item["sheet_name"]))

    return {
        "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "case_dir": str(case_dir.resolve()),
        "workbook_count": len(workbook_paths),
        "workbooks": workbooks,
        "recurring_sheet_names": recurring_sheet_names,
    }


def main():
    parser = argparse.ArgumentParser(description="盘点 cases/ 下 Excel 工作簿的结构特征")
    parser.add_argument(
        "--cases-dir",
        default="cases",
        help="案例工作簿目录，默认读取 cases/",
    )
    parser.add_argument(
        "--output",
        help="可选，输出 JSON 文件路径；不传则打印到标准输出",
    )
    args = parser.parse_args()

    case_dir = Path(args.cases_dir).resolve()
    if not case_dir.exists():
        raise FileNotFoundError(f"案例目录不存在: {case_dir}")

    analysis_result = analyze_cases(case_dir)

    if args.output:
        output_path = Path(args.output).resolve()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as handle:
            json.dump(analysis_result, handle, ensure_ascii=False, indent=2)
            handle.write("\n")
        print(f"[OK] 分析结果已写入: {output_path}")
        return

    print(json.dumps(analysis_result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
