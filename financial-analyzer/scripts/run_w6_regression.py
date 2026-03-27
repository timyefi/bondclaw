#!/usr/bin/env python3
"""
W6.1 回归检查入口。
"""

import argparse
import datetime
import json
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[2]
TEST_RUNS_DIR = ROOT_DIR / "financial-analyzer" / "test_runs"
TESTDATA_DIR = ROOT_DIR / "financial-analyzer" / "testdata"
GOLDEN_DIR = TESTDATA_DIR / "w6_golden"
ANALYZER_SCRIPT = ROOT_DIR / "financial-analyzer" / "scripts" / "financial_analyzer.py"

CORE_SHEETS = [
    "00_overview",
    "01_kpi_dashboard",
    "02_financial_summary",
    "03_debt_profile",
    "04_liquidity_and_covenants",
    "99_evidence_index",
]

REQUIRED_PAYLOAD_KEYS = [
    "overview",
    "kpi_dashboard",
    "financial_summary",
    "debt_profile",
    "liquidity_and_covenants",
    "evidence_index",
]

SCAFFOLD_OUTPUT_FILENAMES = [
    "chapter_records.jsonl",
    "analysis_report_scaffold.md",
    "focus_list_scaffold.json",
    "final_data_scaffold.json",
    "soul_export_payload_scaffold.json",
]

FORMAL_OUTPUT_FILENAMES = [
    "analysis_report.md",
    "final_data.json",
    "soul_export_payload.json",
    "financial_output.xlsx",
]

PREVIEW_OUTPUT_FILENAMES = [
    "preview.pdf",
]

CASES = {
    "henglong": {
        "label": "恒隆地产",
        "expected_outcome": "success",
        "expected_returncode": 0,
        "md_path": ROOT_DIR / "output" / "恒隆地产" / "恒隆地产2024年報" / "恒隆地产2024年報.md",
        "notes_workfile": TESTDATA_DIR / "henglong_notes_workfile.json",
        "run_dir": TEST_RUNS_DIR / "w6_henglong",
        "golden_path": GOLDEN_DIR / "henglong_success.json",
    },
    "country_garden": {
        "label": "碧桂园",
        "expected_outcome": "success",
        "expected_returncode": 0,
        "md_path": ROOT_DIR / "cases" / "碧桂园2024年年报分析.md",
        "notes_workfile": TESTDATA_DIR / "country_garden_notes_workfile.json",
        "run_dir": TEST_RUNS_DIR / "w6_country_garden",
        "golden_path": GOLDEN_DIR / "country_garden_success.json",
    },
    "hanghai": {
        "label": "杭海新城控股",
        "expected_outcome": "success",
        "expected_returncode": 0,
        "md_path": ROOT_DIR / "cases" / "杭海新城控股2024年年报分析.md",
        "notes_workfile": TESTDATA_DIR / "hanghai_notes_workfile.json",
        "run_dir": TEST_RUNS_DIR / "w6_hanghai",
        "golden_path": GOLDEN_DIR / "hanghai_success.json",
    },
    "missing_notes_workfile": {
        "label": "恒隆地产 / 缺失 notes_workfile",
        "expected_outcome": "failure",
        "expected_returncode": 1,
        "expected_failure_reason": "notes_workfile_missing",
        "md_path": ROOT_DIR / "output" / "恒隆地产" / "恒隆地产2024年報" / "恒隆地产2024年報.md",
        "notes_workfile": TESTDATA_DIR / "missing_notes_workfile.json",
        "run_dir": TEST_RUNS_DIR / "w6_missing_notes_workfile",
        "golden_path": GOLDEN_DIR / "missing_notes_workfile_failure.json",
    },
    "invalid_notes_workfile": {
        "label": "恒隆地产 / 无效 notes_workfile",
        "expected_outcome": "failure",
        "expected_returncode": 1,
        "expected_failure_reason": "notes_workfile_invalid",
        "md_path": ROOT_DIR / "output" / "恒隆地产" / "恒隆地产2024年報" / "恒隆地产2024年報.md",
        "notes_workfile": TESTDATA_DIR / "henglong_notes_workfile_invalid.json",
        "run_dir": TEST_RUNS_DIR / "w6_invalid_notes_workfile",
        "golden_path": GOLDEN_DIR / "invalid_notes_workfile_failure.json",
    },
}


def now_iso():
    return datetime.datetime.now().astimezone().isoformat(timespec="seconds")


def read_json(path):
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def write_json(path, payload):
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)


def read_text(path):
    with open(path, "r", encoding="utf-8") as handle:
        return handle.read()


def tail_text(text, line_limit=40):
    lines = text.strip().splitlines()
    if len(lines) <= line_limit:
        return "\n".join(lines)
    return "\n".join(lines[-line_limit:])


def make_check(name, passed, details=None, errors=None):
    return {
        "name": name,
        "passed": passed,
        "details": details or {},
        "errors": errors or [],
    }


def make_evaluation(name, status, details=None, diffs=None):
    return {
        "name": name,
        "status": status,
        "details": details or {},
        "diffs": diffs or [],
    }


def normalize_path(value):
    return str(Path(value).resolve())


def build_command(case_config):
    return [
        sys.executable,
        str(ANALYZER_SCRIPT),
        "--md",
        str(case_config["md_path"]),
        "--notes-workfile",
        str(case_config["notes_workfile"]),
        "--run-dir",
        str(case_config["run_dir"]),
    ]


def validate_success_manifest(run_dir):
    manifest_path = run_dir / "run_manifest.json"
    errors = []
    details = {
        "path": str(manifest_path),
        "exists": manifest_path.exists(),
    }
    manifest = None

    if not manifest_path.exists():
        errors.append("run_manifest.json 未生成")
        return make_check("run_manifest", False, details, errors), manifest

    manifest = read_json(manifest_path)
    details["status"] = manifest.get("status")
    if manifest.get("status") != "success":
        errors.append(f"status 不是 success: {manifest.get('status')!r}")

    details["script_output_mode"] = manifest.get("script_output_mode")
    details["codex_review_required"] = manifest.get("codex_review_required")
    if manifest.get("script_output_mode") != "scaffold_only":
        errors.append(f"script_output_mode 不是 scaffold_only: {manifest.get('script_output_mode')!r}")
    if manifest.get("codex_review_required") is not True:
        errors.append("codex_review_required 不是 true")

    artifacts = manifest.get("artifacts") or {}
    required_artifacts = {
        "run_manifest": run_dir / "run_manifest.json",
        "chapter_records": run_dir / "chapter_records.jsonl",
        "analysis_report_scaffold": run_dir / "analysis_report_scaffold.md",
        "focus_list_scaffold": run_dir / "focus_list_scaffold.json",
        "final_data_scaffold": run_dir / "final_data_scaffold.json",
        "soul_export_payload_scaffold": run_dir / "soul_export_payload_scaffold.json",
    }
    artifact_details = {}

    for key, expected_path in required_artifacts.items():
        actual_value = artifacts.get(key)
        entry = {
            "expected": str(expected_path.resolve()),
            "actual": actual_value,
            "exists": False,
            "matches_run_dir": False,
        }
        if not actual_value:
            errors.append(f"artifacts.{key} 缺失")
            artifact_details[key] = entry
            continue

        actual_path = Path(actual_value)
        entry["exists"] = actual_path.exists()
        entry["matches_run_dir"] = normalize_path(actual_value) == str(expected_path.resolve())
        artifact_details[key] = entry

        if not entry["exists"]:
            errors.append(f"artifacts.{key} 指向文件不存在: {actual_value}")
        if not entry["matches_run_dir"]:
            errors.append(f"artifacts.{key} 未指向本次 run dir: {actual_value}")

    details["artifacts"] = artifact_details
    return make_check("run_manifest", not errors, details, errors), manifest


def validate_failure_manifest(run_dir, expected_failure_reason):
    manifest_path = run_dir / "run_manifest.json"
    errors = []
    details = {
        "path": str(manifest_path),
        "exists": manifest_path.exists(),
        "expected_failure_reason": expected_failure_reason,
    }
    manifest = None

    if not manifest_path.exists():
        errors.append("run_manifest.json 未生成")
        return make_check("run_manifest_failure", False, details, errors), manifest

    manifest = read_json(manifest_path)
    details["status"] = manifest.get("status")
    details["failure_reason"] = manifest.get("failure_reason")
    details["notes_locator_status"] = ((manifest.get("notes_locator") or {}).get("status"))
    artifact_names = sorted((manifest.get("artifacts") or {}).keys())
    details["artifact_names"] = artifact_names

    if manifest.get("status") != "failed":
        errors.append(f"status 不是 failed: {manifest.get('status')!r}")
    if manifest.get("failure_reason") != expected_failure_reason:
        errors.append(
            "failure_reason 不匹配: "
            f"expected={expected_failure_reason!r}, actual={manifest.get('failure_reason')!r}"
        )
    if details["notes_locator_status"] != "failed":
        errors.append(
            "notes_locator.status 不匹配: "
            f"expected='failed', actual={details['notes_locator_status']!r}"
        )

    artifacts = manifest.get("artifacts") or {}
    run_manifest_artifact = artifacts.get("run_manifest")
    expected_run_manifest_path = str((run_dir / "run_manifest.json").resolve())
    details["run_manifest_artifact"] = run_manifest_artifact
    if run_manifest_artifact != expected_run_manifest_path:
        errors.append(
            "artifacts.run_manifest 未指向本次 run dir: "
            f"{run_manifest_artifact!r}"
        )

    existing_artifact_names = []
    for name, artifact_path in artifacts.items():
        if Path(artifact_path).exists():
            existing_artifact_names.append(name)
    details["existing_artifact_names"] = sorted(existing_artifact_names)
    if sorted(existing_artifact_names) != ["run_manifest"]:
        errors.append(
            "失败态实际存在的 artifact 不符合预期: "
            + ", ".join(sorted(existing_artifact_names))
        )

    return make_check("run_manifest_failure", not errors, details, errors), manifest


def validate_analysis_report(run_dir):
    report_path = run_dir / "analysis_report_scaffold.md"
    errors = []
    details = {
        "path": str(report_path),
        "exists": report_path.exists(),
    }

    if not report_path.exists():
        errors.append("analysis_report_scaffold.md 未生成")
        return make_check("analysis_report", False, details, errors)

    content = read_text(report_path)
    details["size_bytes"] = report_path.stat().st_size
    details["contains_run_overview"] = "运行概览" in content
    details["contains_dynamic_focus"] = "动态重点" in content
    details["contains_codex_checklist"] = "Codex 复核清单" in content

    if not content.strip():
        errors.append("analysis_report_scaffold.md 为空")
    if not details["contains_run_overview"]:
        errors.append("analysis_report_scaffold.md 缺少“运行概览”")
    if not details["contains_dynamic_focus"]:
        errors.append("analysis_report_scaffold.md 缺少“脚本初步重点/动态重点”")
    if not details["contains_codex_checklist"]:
        errors.append("analysis_report_scaffold.md 缺少“Codex 复核清单”")

    return make_check("analysis_report_scaffold", not errors, details, errors)


def validate_payload(run_dir):
    payload_path = run_dir / "soul_export_payload_scaffold.json"
    errors = []
    details = {
        "path": str(payload_path),
        "exists": payload_path.exists(),
    }
    payload = None

    if not payload_path.exists():
        errors.append("soul_export_payload_scaffold.json 未生成")
        return make_check("soul_export_payload", False, details, errors), payload

    payload = read_json(payload_path)
    details["contract_version"] = payload.get("contract_version")
    details["template_version"] = payload.get("template_version")
    details["has_required_keys"] = {key: key in payload for key in REQUIRED_PAYLOAD_KEYS}
    details["module_manifest_count"] = len(payload.get("module_manifest") or [])
    details["evidence_index_count"] = len(payload.get("evidence_index") or [])

    if payload.get("contract_version") != "soul_export_v1":
        errors.append(f"contract_version 非预期: {payload.get('contract_version')!r}")
    if payload.get("template_version") != "soul_v1_1_alpha":
        errors.append(f"template_version 非预期: {payload.get('template_version')!r}")

    for key in REQUIRED_PAYLOAD_KEYS:
        if key not in payload:
            errors.append(f"payload 缺少固定骨架键: {key}")

    if not payload.get("module_manifest"):
        errors.append("module_manifest 为空")

    return make_check("soul_export_payload_scaffold", not errors, details, errors), payload

def validate_scaffold_outputs_present(run_dir):
    errors = []
    details = {"checked_paths": {}}
    for filename in SCAFFOLD_OUTPUT_FILENAMES:
        path = run_dir / filename
        exists = path.exists()
        size = path.stat().st_size if exists else 0
        details["checked_paths"][filename] = {
            "path": str(path),
            "exists": exists,
            "size_bytes": size,
        }
        if not exists:
            errors.append(f"成功态缺少 scaffold 文件: {filename}")
        elif size == 0:
            errors.append(f"成功态 scaffold 文件为空: {filename}")
    return make_check("scaffold_outputs_present", not errors, details, errors)


def validate_formal_outputs_absent(run_dir):
    errors = []
    details = {"checked_paths": {}}
    all_absent = FORMAL_OUTPUT_FILENAMES + PREVIEW_OUTPUT_FILENAMES
    for filename in all_absent:
        path = run_dir / filename
        exists = path.exists()
        details["checked_paths"][filename] = {
            "path": str(path),
            "exists": exists,
        }
        if exists:
            errors.append(f"模板脚本不应生成正式产物: {filename}")

    preview_pngs = sorted(run_dir.glob("preview-*.png"))
    details["preview_png_count"] = len(preview_pngs)
    details["preview_png_files"] = [path.name for path in preview_pngs]
    if preview_pngs:
        errors.append("模板脚本不应生成 preview-*.png")

    return make_check("formal_outputs_absent", not errors, details, errors)


def validate_failure_outputs_absent(run_dir):
    errors = []
    details = {"checked_paths": {}}
    all_absent = SCAFFOLD_OUTPUT_FILENAMES + FORMAL_OUTPUT_FILENAMES + PREVIEW_OUTPUT_FILENAMES
    for filename in all_absent:
        path = run_dir / filename
        exists = path.exists()
        details["checked_paths"][filename] = {
            "path": str(path),
            "exists": exists,
        }
        if exists:
            errors.append(f"失败态不应生成任何成功态产物: {filename}")

    preview_pngs = sorted(run_dir.glob("preview-*.png"))
    details["preview_png_count"] = len(preview_pngs)
    details["preview_png_files"] = [path.name for path in preview_pngs]
    if preview_pngs:
        errors.append("失败态不应生成 preview-*.png")

    return make_check("failure_outputs_absent", not errors, details, errors)


def build_success_golden_subset(payload):
    entity_profile = payload.get("entity_profile") or {}
    return {
        "entity_profile": {
            key: entity_profile.get(key)
            for key in [
                "company_name",
                "report_period",
                "currency",
                "report_type",
                "audit_opinion",
                "industry_tag",
            ]
        },
        "module_manifest": [
            {key: item.get(key) for key in ["module_key", "sheet_name", "enabled"]}
            for item in payload.get("module_manifest") or []
        ],
        "overview": {
            "key_risks": [
                {key: item.get(key) for key in ["risk_code", "risk_level"]}
                for item in (payload.get("overview") or {}).get("key_risks") or []
            ]
        },
        "debt_profile": {
            "totals": [
                {key: item.get(key) for key in ["metric_code", "value", "unit", "source_status"]}
                for item in (payload.get("debt_profile") or {}).get("totals") or []
            ]
        },
        "liquidity_and_covenants": {
            "cash_metrics": [
                {key: item.get(key) for key in ["metric_code", "value", "unit", "source_status"]}
                for item in (payload.get("liquidity_and_covenants") or {}).get("cash_metrics") or []
            ]
        },
    }


def build_failure_golden_subset(manifest):
    notes_locator = manifest.get("notes_locator") or {}
    notes_catalog_summary = manifest.get("notes_catalog_summary") or {}
    return {
        "status": manifest.get("status"),
        "failure_reason": manifest.get("failure_reason"),
        "notes_locator": {
            "status": notes_locator.get("status"),
        },
        "notes_catalog_summary": {
            "note_chapter_count": notes_catalog_summary.get("note_chapter_count"),
        },
        "existing_artifacts": sorted(
            name
            for name, artifact_path in (manifest.get("artifacts") or {}).items()
            if Path(artifact_path).exists()
        ),
    }


def collect_diffs(expected, actual, path, diffs):
    if isinstance(expected, dict) and isinstance(actual, dict):
        all_keys = sorted(set(expected.keys()) | set(actual.keys()))
        for key in all_keys:
            next_path = f"{path}.{key}" if path else key
            if key not in expected:
                diffs.append(
                    f"{next_path}: expected=<missing>, actual={json.dumps(actual[key], ensure_ascii=False)}"
                )
                continue
            if key not in actual:
                diffs.append(
                    f"{next_path}: expected={json.dumps(expected[key], ensure_ascii=False)}, actual=<missing>"
                )
                continue
            collect_diffs(expected[key], actual[key], next_path, diffs)
        return

    if isinstance(expected, list) and isinstance(actual, list):
        if len(expected) != len(actual):
            diffs.append(f"{path}: expected_len={len(expected)}, actual_len={len(actual)}")
        for index, (expected_item, actual_item) in enumerate(zip(expected, actual)):
            next_path = f"{path}[{index}]"
            collect_diffs(expected_item, actual_item, next_path, diffs)
        return

    if expected != actual:
        diffs.append(
            f"{path}: expected={json.dumps(expected, ensure_ascii=False)}, "
            f"actual={json.dumps(actual, ensure_ascii=False)}"
        )


def evaluate_golden(case_config, actual_subset, evaluation_name):
    golden_path = case_config["golden_path"]
    details = {
        "golden_path": str(golden_path),
        "golden_exists": golden_path.exists(),
    }
    if not golden_path.exists():
        return make_evaluation(evaluation_name, "skipped", details, ["golden 基线文件不存在"])

    expected_subset = read_json(golden_path)
    diffs = []
    collect_diffs(expected_subset, actual_subset, "", diffs)
    details["diff_count"] = len(diffs)
    return make_evaluation(
        evaluation_name,
        "clean" if not diffs else "diff",
        details,
        diffs,
    )


def run_case(case_id, case_config):
    run_dir = case_config["run_dir"]
    if run_dir.exists():
        shutil.rmtree(run_dir)
    run_dir.mkdir(parents=True, exist_ok=True)

    command = build_command(case_config)
    started_at = now_iso()
    completed = subprocess.run(
        command,
        cwd=str(ROOT_DIR),
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    finished_at = now_iso()

    checks = []
    evaluations = []
    if case_config["expected_outcome"] == "success":
        manifest_check, manifest = validate_success_manifest(run_dir)
        scaffold_check = validate_scaffold_outputs_present(run_dir)
        report_check = validate_analysis_report(run_dir)
        payload_check, payload = validate_payload(run_dir)
        formal_absent_check = validate_formal_outputs_absent(run_dir)
        checks = [manifest_check, scaffold_check, report_check, payload_check, formal_absent_check]
        if payload is not None:
            evaluations.append(
                evaluate_golden(
                    case_config,
                    build_success_golden_subset(payload),
                    "golden_success_subset",
                )
            )
        else:
            evaluations.append(
                make_evaluation(
                    "golden_success_subset",
                    "skipped",
                    {"golden_path": str(case_config["golden_path"])},
                    ["payload 缺失，无法执行 golden diff"],
                )
            )
    else:
        manifest_check, manifest = validate_failure_manifest(
            run_dir,
            case_config["expected_failure_reason"],
        )
        absent_check = validate_formal_outputs_absent(run_dir)
        scaffold_absent_check = validate_failure_outputs_absent(run_dir)
        checks = [manifest_check, absent_check, scaffold_absent_check]
        if manifest is not None:
            evaluations.append(
                evaluate_golden(
                    case_config,
                    build_failure_golden_subset(manifest),
                    "golden_failure_subset",
                )
            )
        else:
            evaluations.append(
                make_evaluation(
                    "golden_failure_subset",
                    "skipped",
                    {"golden_path": str(case_config["golden_path"])},
                    ["run_manifest 缺失，无法执行 golden diff"],
                )
            )

    matched = completed.returncode == case_config["expected_returncode"] and all(
        item["passed"] for item in checks
    )

    return {
        "case_id": case_id,
        "label": case_config["label"],
        "expected_outcome": case_config["expected_outcome"],
        "md_path": str(case_config["md_path"]),
        "notes_workfile": str(case_config["notes_workfile"]),
        "run_dir": str(run_dir),
        "golden_path": str(case_config["golden_path"]),
        "started_at": started_at,
        "finished_at": finished_at,
        "expected_returncode": case_config["expected_returncode"],
        "returncode": completed.returncode,
        "matched": matched,
        "passed": matched,
        "command": command,
        "stdout_tail": tail_text(completed.stdout),
        "stderr_tail": tail_text(completed.stderr),
        "checks": checks,
        "evaluations": evaluations,
    }


def load_sheetnames(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def collect_known_gaps():
    gaps = []

    legacy_manifest_cases = [
        ("henglong_notes_only", TEST_RUNS_DIR / "henglong_notes_only" / "run_manifest.json"),
        ("henglong_v3", TEST_RUNS_DIR / "henglong_v3" / "run_manifest.json"),
    ]
    windows_path_examples = []
    for case_id, manifest_path in legacy_manifest_cases:
        if not manifest_path.exists():
            continue
        manifest = read_json(manifest_path)
        artifact_path = ((manifest.get("artifacts") or {}).get("run_manifest")) or ""
        if artifact_path.startswith("C:\\"):
            windows_path_examples.append({
                "case_id": case_id,
                "path": artifact_path,
            })

    if windows_path_examples:
        gaps.append({
            "id": "legacy_windows_artifact_paths",
            "category": "observed_repo_state",
            "title": "历史 run_manifest 仍保留 Windows 绝对路径",
            "detail": "历史 test_runs 样本不能直接作为当前 W6 基线，因为 artifacts 仍指向旧环境路径。",
            "evidence": windows_path_examples,
        })

    henglong_v3_manifest = TEST_RUNS_DIR / "henglong_v3" / "run_manifest.json"
    if henglong_v3_manifest.exists():
        manifest = read_json(henglong_v3_manifest)
        if manifest.get("status") is None:
            gaps.append({
                "id": "legacy_manifest_missing_status",
                "category": "observed_repo_state",
                "title": "henglong_v3 缺少正式 status 字段",
                "detail": "该目录可作为历史排障样本，但不能继续充当成功态回归基线。",
                "evidence": [
                    {
                        "case_id": "henglong_v3",
                        "manifest_path": str(henglong_v3_manifest),
                    }
                ],
            })

    legacy_workbook = TEST_RUNS_DIR / "henglong_soul_contract" / "financial_output.xlsx"
    soul_workbook = TEST_RUNS_DIR / "henglong_soul_v1_1_alpha" / "soul_output.xlsx"
    if legacy_workbook.exists() and soul_workbook.exists():
        legacy_sheetnames = load_sheetnames(legacy_workbook)
        soul_sheetnames = load_sheetnames(soul_workbook)
        if "summary" in legacy_sheetnames and "00_overview" in soul_sheetnames:
            gaps.append({
                "id": "mixed_legacy_and_soul_workbooks",
                "category": "observed_repo_state",
                "title": "历史目录同时存在旧内部 workbook 与新 Soul workbook",
                "detail": "历史目录仍混有旧 workbook 样本；当前 W6 已切换为 scaffold-only，不再用旧 workbook 作为模板脚本成功依据。",
                "evidence": [
                    {
                        "path": str(legacy_workbook),
                        "sheetnames": legacy_sheetnames,
                    },
                    {
                        "path": str(soul_workbook),
                        "sheetnames": soul_sheetnames,
                    },
                ],
            })

    country_garden_workbook = TEST_RUNS_DIR / "w6_country_garden" / "financial_output.xlsx"
    if country_garden_workbook.exists():
        gaps.append({
            "id": "workbook_cell_level_golden_deferred",
            "category": "known_scope_boundary",
            "title": "Workbook 单元格级 golden diff 仍暂缓",
            "detail": "当前 W6.1 仅冻结 scaffold 产物和受控 payload 子集；逐单元格 workbook golden 已移出模板脚本门禁，留待正式导出层或单独 QA 轨。",
            "evidence": [
                {
                    "path": str(country_garden_workbook),
                    "sheet_name": "01_kpi_dashboard",
                    "worksheet_xml": "xl/worksheets/sheet2.xml",
                }
            ],
        })

    gaps.append({
        "id": "preview_layout_semantics_not_covered",
        "category": "known_scope_boundary",
        "title": "模板脚本不再把 preview 作为主门禁",
        "detail": "当前模板脚本只负责 scaffold 产物；preview.pdf / preview-*.png 已从 W6 成功态门禁中移出，不再作为模板脚本验收项。",
        "evidence": [],
    })

    return gaps


def build_summary(results):
    evaluation_status_counts = {
        "clean": 0,
        "diff": 0,
        "skipped": 0,
    }
    for result in results:
        for evaluation in result["evaluations"]:
            status = evaluation["status"]
            if status in evaluation_status_counts:
                evaluation_status_counts[status] += 1

    expected_success_count = sum(1 for item in results if item["expected_outcome"] == "success")
    expected_failure_count = sum(1 for item in results if item["expected_outcome"] == "failure")
    matched_case_count = sum(1 for item in results if item["matched"])
    mismatched_case_count = len(results) - matched_case_count
    return {
        "selected_case_count": len(results),
        "expected_success_count": expected_success_count,
        "expected_failure_count": expected_failure_count,
        "matched_case_count": matched_case_count,
        "mismatched_case_count": mismatched_case_count,
        "all_passed": mismatched_case_count == 0,
        "evaluation_status_counts": evaluation_status_counts,
    }


def render_case_block(result):
    status_text = "MATCHED" if result["matched"] else "MISMATCHED"
    lines = [
        f"### {result['label']} (`{result['case_id']}`) - {status_text}",
        "",
        f"- expected_outcome: `{result['expected_outcome']}`",
        f"- run dir: `{result['run_dir']}`",
        f"- markdown: `{result['md_path']}`",
        f"- notes_workfile: `{result['notes_workfile']}`",
        f"- expected_returncode: `{result['expected_returncode']}`",
        f"- returncode: `{result['returncode']}`",
        "",
        "#### Checks",
        "",
    ]

    for check in result["checks"]:
        marker = "PASS" if check["passed"] else "FAIL"
        lines.append(f"- `{check['name']}`: {marker}")
        for error in check["errors"]:
            lines.append(f"  - {error}")

    if result["stdout_tail"]:
        lines.extend([
            "",
            "```text",
            result["stdout_tail"],
            "```",
        ])

    if result["stderr_tail"]:
        lines.extend([
            "",
            "```text",
            result["stderr_tail"],
            "```",
        ])

    lines.append("")
    return lines


def render_report(results_payload):
    summary = results_payload["summary"]
    success_results = [
        result for result in results_payload["results"] if result["expected_outcome"] == "success"
    ]
    failure_results = [
        result for result in results_payload["results"] if result["expected_outcome"] == "failure"
    ]
    lines = [
        "# W6 Regression Report",
        "",
        f"- 生成时间：{results_payload['generated_at']}",
        f"- 结果 JSON：`{results_payload['results_path']}`",
        f"- 匹配情况：{summary['matched_case_count']}/{summary['selected_case_count']} matched",
        f"- 成功案例：{summary['expected_success_count']}",
        f"- 失败案例：{summary['expected_failure_count']}",
        f"- Golden Diff：clean={summary['evaluation_status_counts']['clean']}, diff={summary['evaluation_status_counts']['diff']}, skipped={summary['evaluation_status_counts']['skipped']}",
        "",
        "## 验证范围",
        "",
        "- 固定 3 个成功案例重跑 `financial_analyzer.py`，并验证 `run_manifest.json`、`chapter_records.jsonl`、`analysis_report_scaffold.md`、`focus_list_scaffold.json`、`final_data_scaffold.json`、`soul_export_payload_scaffold.json`。",
        "- 固定 3 个成功案例还需确认 `analysis_report.md`、`final_data.json`、`soul_export_payload.json`、`financial_output.xlsx`、`preview.pdf`、`preview-*.png` 均未生成。",
        "- 固定 2 个失败案例重跑 `financial_analyzer.py`，并验证失败态 `run_manifest.json`、失败原因和所有 scaffold / 正式产物缺失。",
        "- Golden diff 仅评估受控 payload/manifest 子集，不把 diff 作为退出码门禁。",
        "- 不把历史 `test_runs` 目录作为通过依据。",
        "",
        "## 成功案例",
        "",
    ]

    for result in success_results:
        lines.extend(render_case_block(result))

    lines.extend([
        "## 失败路径回归",
        "",
    ])
    for result in failure_results:
        lines.extend(render_case_block(result))

    lines.extend([
        "## Golden Diff 评估",
        "",
    ])
    for result in results_payload["results"]:
        lines.append(f"### {result['label']} (`{result['case_id']}`)")
        lines.append("")
        for evaluation in result["evaluations"]:
            lines.append(f"- `{evaluation['name']}`: `{evaluation['status']}`")
            if evaluation["details"].get("golden_path"):
                lines.append(f"  - golden_path: `{evaluation['details']['golden_path']}`")
            if evaluation["diffs"]:
                lines.append("  - diffs:")
                for diff in evaluation["diffs"]:
                    lines.append(f"    - {diff}")
        lines.append("")

    lines.extend([
        "## 已知缺口",
        "",
    ])

    for gap in results_payload["known_gaps"]:
        lines.append(f"### {gap['title']}")
        lines.append("")
        lines.append(f"- 分类：`{gap['category']}`")
        lines.append(f"- 说明：{gap['detail']}")
        if gap["evidence"]:
            lines.append("- 证据：")
            for entry in gap["evidence"]:
                lines.append(f"  - `{json.dumps(entry, ensure_ascii=False)}`")
        lines.append("")

    return "\n".join(lines).rstrip() + "\n"


def parse_args():
    parser = argparse.ArgumentParser(description="运行 W6.1 回归检查")
    parser.add_argument(
        "--case",
        action="append",
        choices=sorted(CASES.keys()),
        help="仅运行指定案例；可重复传入多个 --case",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    selected_case_ids = args.case or list(CASES.keys())
    results = [run_case(case_id, CASES[case_id]) for case_id in selected_case_ids]
    summary = build_summary(results)
    known_gaps = collect_known_gaps()

    results_path = TEST_RUNS_DIR / "w6_regression_results.json"
    report_path = TEST_RUNS_DIR / "w6_regression_report.md"
    payload = {
        "generated_at": now_iso(),
        "script": str(Path(__file__).resolve()),
        "selected_cases": selected_case_ids,
        "summary": summary,
        "results": results,
        "known_gaps": known_gaps,
        "results_path": str(results_path),
        "report_path": str(report_path),
    }

    write_json(results_path, payload)
    report = render_report(payload)
    with open(report_path, "w", encoding="utf-8") as handle:
        handle.write(report)

    print(f"[INFO] 已写出结果 JSON: {results_path}")
    print(f"[INFO] 已写出结果报告: {report_path}")
    print(
        "[OK] 回归结果: "
        f"{summary['matched_case_count']}/{summary['selected_case_count']} matched"
    )

    raise SystemExit(0 if summary["all_passed"] else 1)


if __name__ == "__main__":
    main()
