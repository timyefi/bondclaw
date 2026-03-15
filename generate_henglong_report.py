from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# 删除默认sheet
wb.remove(wb.active)

# 1. 封面
ws_cover = wb.create_sheet('封面')
ws_cover['B2'] = '恒隆地产有限公司'
ws_cover['B2'].font = Font(size=24, bold=True)
ws_cover['B3'] = '2024年度财务分析报告'
ws_cover['B3'].font = Font(size=18)
ws_cover['B5'] = '报告日期：2026年3月14日'
ws_cover['B6'] = '分析对象：恒隆地产2024年年报（港股）'
ws_cover['B8'] = '重要说明：'
ws_cover['B9'] = '• 本报告基于恒隆地产2024年年报附注逐章分析'
ws_cover['B10'] = '• 采用HKFRS（香港财务报告准则）'
ws_cover['B11'] = '• 投资物业按公允价值计量'

# 2. 核心财务指标
ws_core = wb.create_sheet('核心财务指标')
headers = ['指标类别', '指标名称', '2024年', '2023年', '变动', '单位']
for col, header in enumerate(headers, 1):
    ws_core.cell(1, col, header).font = Font(bold=True)
    ws_core.cell(1, col).fill = PatternFill('solid', start_color='D9E1F2')

data = [
    ['收入', '总收入', 112.42, 103.16, '8.97%', '亿港元'],
    ['', '租赁收入', 103.00, 95.00, '8.42%', '亿港元'],
    ['', '物业销售收入', 1.89, 1.54, '22.73%', '亿港元'],
    ['盈利', '股东应占溢利', 22.20, 28.11, '-21.03%', '亿港元'],
    ['', '除税前溢利', 47.16, 50.37, '-6.37%', '亿港元'],
    ['', '实际税率', 29.43, 31.21, '-1.78pct', '%'],
    ['资产', '总资产', 2216.48, 2173.02, '2.00%', '亿港元'],
    ['', '投资物业', 1665.19, 1690.46, '-1.49%', '亿港元'],
    ['', '发展中投资物业', 240.01, 236.10, '1.66%', '亿港元'],
    ['', '待售物业', 134.89, 142.23, '-5.16%', '亿港元'],
    ['负债', '有息债务合计', 575.74, 508.98, '13.12%', '亿港元'],
    ['', '银行贷款', 426.54, 372.35, '14.55%', '亿港元'],
    ['', '债券', 149.20, 136.63, '9.20%', '亿港元'],
    ['', '一年内到期债务', 93.40, 44.34, '110.65%', '亿港元'],
    ['现金', '现金及银行存款', 103.03, 53.52, '92.51%', '亿港元'],
    ['', '现金及现金等价物', 101.98, 48.91, '108.51%', '亿港元'],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_core.cell(row_idx, col_idx, value)
        if col_idx in [3, 4] and isinstance(value, (int, float)):
            cell.number_format = '#,##0.00'

# 3. 偿债能力分析
ws_solvency = wb.create_sheet('偿债能力分析')
headers = ['指标名称', '计算公式', '2024年', '2023年', '评估标准', '评估结果']
for col, header in enumerate(headers, 1):
    ws_solvency.cell(1, col, header).font = Font(bold=True)
    ws_solvency.cell(1, col).fill = PatternFill('solid', start_color='D9E1F2')

solvency_data = [
    ['流动比率', '流动资产/流动负债', 1.42, 1.52, '>1.0', '健康'],
    ['速动比率', '(流动资产-存货)/流动负债', 0.71, 0.58, '>0.5', '健康'],
    ['现金比率', '现金/流动负债', 0.54, 0.32, '>0.2', '健康'],
    ['资产负债率', '总负债/总资产', 36.37, 34.57, '<70%', '健康'],
    ['净负债率', '净债项/总资产', 21.24, 20.87, '<50%', '健康'],
    ['现金短债比', '现金/一年内到期债务', 1.09, 1.10, '>1.0', '健康'],
    ['利息保障倍数', 'EBIT/利息支出', 4.05, 4.17, '>3.0', '健康'],
]

for row_idx, row_data in enumerate(solvency_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_solvency.cell(row_idx, col_idx, value)
        if col_idx in [3, 4] and isinstance(value, (int, float)):
            cell.number_format = '0.00'

# 4. 有息债务分析
ws_debt = wb.create_sheet('有息债务分析')
headers = ['债务类型', '2024年金额', '2023年金额', '变动', '利率范围', '占比']
for col, header in enumerate(headers, 1):
    ws_debt.cell(1, col, header).font = Font(bold=True)
    ws_debt.cell(1, col).fill = PatternFill('solid', start_color='D9E1F2')

debt_data = [
    ['银行贷款-一年内到期', 44.34, 37.24, '19.09%', '3.4%-7.0%', '7.70%'],
    ['银行贷款-一年后到期', 382.20, 335.11, '14.05%', '3.4%-7.0%', '66.38%'],
    ['银行贷款小计', 426.54, 372.35, '14.55%', '-', '74.09%'],
    ['债券-一年内到期', 49.06, 7.10, '590.99%', '2.0%-5.0%', '8.52%'],
    ['债券-一年后到期', 100.14, 129.53, '-22.69%', '2.0%-5.0%', '17.39%'],
    ['债券小计', 149.20, 136.63, '9.20%', '-', '25.91%'],
    ['有息债务合计', 575.74, 508.98, '13.12%', '-', '100.00%'],
    ['减：一年内到期部分', 93.40, 44.34, '110.65%', '-', '-'],
    ['非流动有息债务', 482.34, 464.64, '3.81%', '-', '-'],
]

for row_idx, row_data in enumerate(debt_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_debt.cell(row_idx, col_idx, value)
        if col_idx in [2, 3] and isinstance(value, (int, float)):
            cell.number_format = '#,##0.00'

# 5. 债务期限结构
ws_maturity = wb.create_sheet('债务期限结构')
ws_maturity['A1'] = '债务期限结构分析（2024年12月31日）'
ws_maturity['A1'].font = Font(bold=True, size=14)

headers = ['期限', '金额（亿港元）', '占比', '累计占比']
for col, header in enumerate(headers, 1):
    ws_maturity.cell(3, col, header).font = Font(bold=True)
    ws_maturity.cell(3, col).fill = PatternFill('solid', start_color='D9E1F2')

maturity_data = [
    ['一年内到期', 93.40, '16.22%', '16.22%'],
    ['一年后至二年内', 73.58, '12.78%', '29.00%'],
    ['二年后至五年内', 167.93, '29.17%', '58.17%'],
    ['五年以上', 240.83, '41.83%', '100.00%'],
]

for row_idx, row_data in enumerate(maturity_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_maturity.cell(row_idx, col_idx, value)
        if col_idx == 2 and isinstance(value, (int, float)):
            cell.number_format = '#,##0.00'

ws_maturity['A9'] = '关键发现：'
ws_maturity['A10'] = '• 短期债务（一年内）占比16.22%，现金覆盖充足（现金短债比1.09x）'
ws_maturity['A11'] = '• 中期债务（1-5年）占比41.95%，期限结构合理'
ws_maturity['A12'] = '• 长期债务（5年以上）占比41.83%，债务期限较长，偿债压力分散'

# 6. 投资物业分析
ws_property = wb.create_sheet('投资物业分析')
ws_property['A1'] = '投资物业公允价值分析'
ws_property['A1'].font = Font(bold=True, size=14)

headers = ['项目', '2024年', '2023年', '变动', '公允价值层级', '估值方法']
for col, header in enumerate(headers, 1):
    ws_property.cell(3, col, header).font = Font(bold=True)
    ws_property.cell(3, col).fill = PatternFill('solid', start_color='D9E1F2')

property_data = [
    ['投资物业-香港长期地契', 366.12, 369.73, '-0.98%', '第二级', '收入资本化法'],
    ['投资物业-香港中期地契', 258.64, 248.54, '4.06%', '第二级', '收入资本化法'],
    ['投资物业-内地中期地契', 1040.43, 1072.19, '-2.96%', '第二级', '收入资本化法'],
    ['投资物业小计', 1665.19, 1690.46, '-1.49%', '-', '-'],
    ['发展中投资物业', 240.01, 236.10, '1.66%', '第三级', '直接比较法'],
    ['投资物业合计', 1905.20, 1926.56, '-1.11%', '-', '-'],
]

for row_idx, row_data in enumerate(property_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_property.cell(row_idx, col_idx, value)
        if col_idx in [2, 3] and isinstance(value, (int, float)):
            cell.number_format = '#,##0.00'

ws_property['A11'] = '估值说明：'
ws_property['A12'] = '• 估值机构：第一太平戴维斯估值及专业顾问有限公司'
ws_property['A13'] = '• 第二级公允价值：使用可观察输入数据（市场租金、资本化率）'
ws_property['A14'] = '• 第三级公允价值：使用不可观察输入数据（预计开发成本、边际利润）'

# 7. 风险分析
ws_risk = wb.create_sheet('风险分析')
ws_risk['A1'] = '恒隆地产2024年风险识别'
ws_risk['A1'].font = Font(bold=True, size=14)

headers = ['风险类别', '风险描述', '风险等级', '关注指标', '备注']
for col, header in enumerate(headers, 1):
    ws_risk.cell(3, col, header).font = Font(bold=True)
    ws_risk.cell(3, col).fill = PatternFill('solid', start_color='FFC7CE')

risk_data = [
    ['流动性风险', '一年内到期债务增长110.65%', '中', '现金短债比1.09x', '现金覆盖充足但需关注'],
    ['汇率风险', '人民币资产占比高，港币报表', '中', '汇兑储备-73.83亿', '2024年人民币贬值影响'],
    ['利率风险', '浮动利率债务占比高', '中', '借贷利率3.4%-7.0%', '美联储降息周期有利'],
    ['市场风险', '投资物业公允价值下降1.49%', '低', '投资物业1665亿', '核心资产价值稳定'],
    ['经营风险', '租金收入增长8.42%', '低', '租赁收入103亿', '核心现金流稳定'],
    [' covenant风险', '355亿借贷受财务契约限制', '低', '完全遵守契约', '无违约风险'],
]

for row_idx, row_data in enumerate(risk_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        ws_risk.cell(row_idx, col_idx, value)

# 8. 技能进化记录
ws_evolution = wb.create_sheet('技能进化记录')
ws_evolution['A1'] = '恒隆地产案例分析技能进化记录'
ws_evolution['A1'].font = Font(bold=True, size=14)

headers = ['进化章节', '发现内容', '技能更新', '影响']
for col, header in enumerate(headers, 1):
    ws_evolution.cell(3, col, header).font = Font(bold=True)
    ws_evolution.cell(3, col).fill = PatternFill('solid', start_color='E2EFDA')

evolution_data = [
    ['第1章-会计政策', '港股采用HKFRS准则', '新增hong_kong_stock分析框架', '港股分析能力'],
    ['第1章-投资物业', '投资物业按公允价值计量', '新增公允价值变动分析', '港股房企特有'],
    ['第1章-借贷成本', '利息100%资本化', '更新利息资本化率频谱', '与A股房企一致'],
    ['第2章-收入结构', '租金收入占比92%', '新增纯商业地产运营商模式', '收入结构分析'],
    ['第17章-债务', '短期债务激增110%', '新增短期债务流动性监控', '风险识别能力'],
    ['第17章-债券', '中期票据计划40亿美元', '新增境外债券分析', '跨境融资分析'],
    ['第14章-现金', '现金大幅增长92%', '新增现金覆盖分析', '流动性分析'],
]

for row_idx, row_data in enumerate(evolution_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        ws_evolution.cell(row_idx, col_idx, value)

# 调整列宽
for ws in wb.worksheets:
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18

# 保存文件
output_path = 'C:\\Users\\Administrator\\Desktop\\项目\\信用工作流\\年报训练\\恒隆地产2024年财务分析.xlsx'
wb.save(output_path)
print(f'报告已生成：{output_path}')
